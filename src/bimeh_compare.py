# src/bimeh_compare.py
import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

def read_table(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return pd.read_excel(path, dtype=object)
    elif ext == ".csv":
        return pd.read_csv(path, dtype=object)
    raise ValueError("Unsupported file type")

def parse_number(val):
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float, np.integer, np.floating)):
        return float(val)
    s = str(val).strip()
    if s == "":
        return np.nan
    if "%" in s or "٪" in s:
        s2 = s.replace("%", "").replace("٪", "")
        s2 = re.sub(r"[^\d\.\-]", "", s2)
        try:
            return float(s2)
        except:
            return np.nan
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except:
        return np.nan

def values_equal(a, b, tolerance=0.01):
    # both empty -> equal
    if (pd.isna(a) or a == "") and (pd.isna(b) or b == ""):
        return True
    a_num = parse_number(a)
    b_num = parse_number(b)
    if (not np.isnan(a_num)) and (not np.isnan(b_num)):
        return float(np.isclose(a_num, b_num, atol=tolerance))
    return str(a).strip() == str(b).strip()

def process_files(bimeh_path, baar_path, out_dir="out", tolerance=0.01):
    os.makedirs(out_dir, exist_ok=True)
    bimeh = read_table(bimeh_path)
    baar = read_table(baar_path)

    # ساده‌سازی: فرض نام ستون‌ها دقیق هستند:
    key = "شماره بارنامه"
    val = "ارزش محموله"
    ins = "مبلغ بیمه"
    tax = "درصد مالیات ارزش افزوده"

    # NaMojood
    bimeh[key] = bimeh[key].astype(str).str.strip()
    baar[key] = baar[key].astype(str).str.strip()
    mask_na = ~bimeh[key].isin(set(baar[key]))
    na_df = bimeh[mask_na]
    if len(na_df) > 0:
        na_path = os.path.join(out_dir, "Bimeh_NaMojood.xlsx")
        na_df.to_excel(na_path, index=False)
        wb = load_workbook(na_path)
        ws = wb.active
        for r in range(2, ws.max_row+1):
            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = ORANGE_FILL
        wb.save(na_path)
        bimeh = bimeh[~mask_na].reset_index(drop=True)

    # drop duplicates keep first
    bimeh = bimeh.drop_duplicates(subset=[key], keep="first")
    baar = baar.drop_duplicates(subset=[key], keep="first")

    merged = pd.merge(
        bimeh[[key, val, ins, tax]],
        baar[[key, val, ins, tax]],
        on=key,
        how="inner",
        suffixes=("_Bimeh", "_Baar")
    )

    out_path = os.path.join(out_dir, "Bimeh_Moghayeseh.xlsx")
    merged.to_excel(out_path, index=False)

    wb = load_workbook(out_path)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {col: idx + 1 for idx, col in enumerate(header)}
    mapping = [
        (f"{val}_Bimeh", f"{val}_Baar"),
        (f"{ins}_Bimeh", f"{ins}_Baar"),
        (f"{tax}_Bimeh", f"{tax}_Baar"),
    ]

    for r in range(2, ws.max_row+1):
        all_equal = True
        mismatched = []
        for left_col, right_col in mapping:
            left_val = ws.cell(row=r, column=col_index[left_col]).value
            right_val = ws.cell(row=r, column=col_index[right_col]).value
            if not values_equal(left_val, right_val, tolerance=tolerance):
                all_equal = False
                mismatched.append(col_index[left_col])
                mismatched.append(col_index[right_col])
        if all_equal:
            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = GREEN_FILL
        else:
            for c in range(1, ws.max_column+1):
                ws.cell(row=r, column=c).fill = RED_FILL
            for c in mismatched:
                ws.cell(row=r, column=c).fill = YELLOW_FILL

    wb.save(out_path)
    print("Done. outputs in:", out_dir)
